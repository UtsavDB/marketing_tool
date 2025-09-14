import base64
import os
import sys
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook

# Ensure repository root on path for module imports
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from core.generate_script_json import invoke_openai_with_image_and_pdf


# ---------------------------------------------------------------------------
# Helper utilities for tests
# ---------------------------------------------------------------------------

def _sheet_to_pdf(sheet, pdf_path: str) -> None:
    """Write the contents of an ``openpyxl`` sheet to a minimal PDF file."""
    lines = []
    for row in sheet.iter_rows(values_only=True):
        line = ",".join("" if cell is None else str(cell) for cell in row)
        lines.append(line)
    text = "\n".join(lines)
    pdf_content = f"%PDF-1.4\n{text}\n%%EOF".encode("utf-8")
    with open(pdf_path, "wb") as f:
        f.write(pdf_content)


def _create_tiny_png(path: str) -> None:
    """Create a 1x1 white PNG without requiring Pillow."""
    tiny_png = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNgYAAAAAMAAWgmWQ0AAAAASUVORK5CYII="
    )
    with open(path, "wb") as f:
        f.write(tiny_png)


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

def test_export_excel_sheet_to_pdf(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Hello"
    ws["B1"] = "World"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    loaded = load_workbook(buf)
    sheet = loaded.active

    pdf_path = tmp_path / "sheet.pdf"
    _sheet_to_pdf(sheet, str(pdf_path))

    assert pdf_path.exists()
    assert pdf_path.stat().st_size > 0


def test_invoke_openai_with_image_and_pdf_encodes_files(monkeypatch, tmp_path):
    image_path = tmp_path / "img.png"
    _create_tiny_png(str(image_path))

    # Reuse PDF creation from previous test
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "One"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    sheet = load_workbook(buf).active
    pdf_path = tmp_path / "sheet.pdf"
    _sheet_to_pdf(sheet, str(pdf_path))

    # Set required environment variables
    monkeypatch.setenv("OPENAI_API_KEY", "test")
    monkeypatch.setenv("OPENAI_API_BASE", "https://example.org")
    monkeypatch.setenv("OPENAI_API_VERSION", "2024-05-01")
    monkeypatch.setenv("OPENAI_DEPLOYMENT_NAME", "model")

    captured = {}

    class DummyCompletions:
        def create(self, *, model, messages):
            captured["model"] = model
            captured["messages"] = messages

            class Resp:
                choices = [
                    type("Obj", (), {"message": type("Obj", (), {"content": "ok"})()})
                ]

            return Resp()

    class DummyClient:
        def __init__(self, **kwargs):
            self.chat = type("Obj", (), {"completions": DummyCompletions()})()

    monkeypatch.setattr(
        "core.generate_script_json.AzureOpenAI", lambda **kwargs: DummyClient()
    )

    result = invoke_openai_with_image_and_pdf("hi", str(image_path), str(pdf_path))
    assert result == "ok"

    messages = captured["messages"]
    assert isinstance(messages, list) and messages
    content = messages[0]["content"]
    types = [part["type"] for part in content]
    assert "image_url" in types
    assert "input_pdf" in types

    img_part = next(part for part in content if part["type"] == "image_url")
    assert img_part["image_url"]["url"].startswith("data:image/png;base64,")

    pdf_part = next(part for part in content if part["type"] == "input_pdf")
    assert pdf_part["data"]
    assert pdf_part["mime_type"] == "application/pdf"
